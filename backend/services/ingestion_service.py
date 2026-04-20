"""
services/ingestion_service.py
-------------------------------
Automatic data ingestion pipeline:
PDF / Excel / CSV → parse → AI classify → stage for posting.
Zero manual entry required for standard workflows.
"""

from __future__ import annotations

import io
import re
from datetime import date
from decimal import Decimal
from typing import Optional
import asyncpg
import structlog

log = structlog.get_logger()


class IngestionService:
    """
    Parses uploaded files and feeds transactions into the AI pipeline.
    Supports: bank statements (CSV/Excel/PDF), sales/purchase invoices.
    """

    def __init__(self, db: asyncpg.Pool):
        self.db = db

    # ── Bank Statement Ingestion ──────────────────────────────────────────

    async def ingest_bank_statement(
        self,
        company_id: str,
        bank_account_id: str,
        file_content: bytes,
        filename: str,
        user_id: Optional[str] = None,
    ) -> dict:
        """Parse bank statement and stage transactions. Returns parse summary."""
        ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

        if ext == "csv":
            transactions = self._parse_csv(file_content)
        elif ext in ("xlsx", "xls"):
            transactions = self._parse_excel(file_content)
        elif ext == "pdf":
            transactions = self._parse_pdf_statement(file_content)
        else:
            raise ValueError(f"Unsupported file type: {ext}. Use CSV, Excel, or PDF.")

        if not transactions:
            raise ValueError("No transactions found in the file. Check the format.")

        # Stage in DB
        staged = 0
        duplicates = 0
        async with self.db.acquire() as conn:
            for txn in transactions:
                try:
                    result = await conn.fetchval(
                        """
                        INSERT INTO bank_transactions
                            (bank_account_id, company_id, txn_date, amount, txn_type,
                             narration, reference, balance, status, raw_data)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,'unmatched',$9)
                        ON CONFLICT (bank_account_id, txn_date, amount, narration)
                        DO NOTHING
                        RETURNING id
                        """,
                        bank_account_id, company_id,
                        txn.get("txn_date"), txn.get("amount"),
                        txn.get("txn_type", "debit"), txn.get("narration"),
                        txn.get("reference"), txn.get("balance"),
                        str(txn)
                    )
                    if result:
                        staged += 1
                    else:
                        duplicates += 1
                except Exception as exc:
                    log.warning("transaction_stage_error", error=str(exc), txn=str(txn)[:100])

        return {
            "total_parsed": len(transactions),
            "staged": staged,
            "duplicates_skipped": duplicates,
            "file_type": ext,
            "bank_account_id": bank_account_id,
        }

    # ── Invoice Ingestion ─────────────────────────────────────────────────

    async def ingest_invoice(
        self,
        company_id: str,
        file_content: bytes,
        filename: str,
        invoice_type: str = "sales",   # 'sales' | 'purchase'
        user_id: Optional[str] = None,
    ) -> dict:
        """Extract invoice data from uploaded file and create journal entry."""
        ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

        if ext == "pdf":
            invoice_data = self._extract_invoice_pdf(file_content)
        elif ext in ("xlsx", "xls", "csv"):
            invoice_data = self._extract_invoice_spreadsheet(file_content, ext)
        else:
            raise ValueError(f"Unsupported invoice format: {ext}")

        if not invoice_data:
            raise ValueError("Could not extract invoice data from file")

        # Store raw invoice
        async with self.db.acquire() as conn:
            invoice_id = await conn.fetchval(
                """
                INSERT INTO uploaded_invoices
                    (company_id, invoice_type, invoice_no, party_name,
                     invoice_date, subtotal, cgst, sgst, igst, total,
                     raw_data, status, uploaded_by, uploaded_at)
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,'pending',$12,NOW())
                RETURNING id::text
                """,
                company_id, invoice_type,
                invoice_data.get("invoice_no", ""),
                invoice_data.get("party_name", "Unknown"),
                invoice_data.get("invoice_date", date.today()),
                Decimal(str(invoice_data.get("subtotal", 0))),
                Decimal(str(invoice_data.get("cgst", 0))),
                Decimal(str(invoice_data.get("sgst", 0))),
                Decimal(str(invoice_data.get("igst", 0))),
                Decimal(str(invoice_data.get("total", 0))),
                str(invoice_data),
                user_id,
            )

        return {
            "invoice_id": invoice_id,
            "invoice_data": invoice_data,
            "status": "pending_journal",
        }

    # ── CSV Parser ────────────────────────────────────────────────────────

    def _parse_csv(self, content: bytes) -> list[dict]:
        import csv
        transactions = []
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))

        for row in reader:
            txn = self._normalize_bank_row(row)
            if txn:
                transactions.append(txn)

        return transactions

    def _parse_excel(self, content: bytes) -> list[dict]:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = None
        transactions = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c or "").lower().strip() for c in row]
                continue
            if not any(row):
                continue
            row_dict = dict(zip(headers or [], row))
            txn = self._normalize_bank_row(row_dict)
            if txn:
                transactions.append(txn)

        return transactions

    def _parse_pdf_statement(self, content: bytes) -> list[dict]:
        """Extract transactions from text-based PDF bank statements."""
        try:
            import pdfplumber
            transactions = []
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages:
                    text = page.extract_text() or ""
                    txns = self._extract_txns_from_text(text)
                    transactions.extend(txns)
            return transactions
        except Exception as exc:
            log.warning("pdf_parse_error", error=str(exc))
            return []

    def _extract_invoice_pdf(self, content: bytes) -> dict:
        """Extract invoice fields from PDF using pattern matching."""
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                text = "\n".join(page.extract_text() or "" for page in pdf.pages)

            return self._parse_invoice_text(text)
        except Exception as exc:
            log.warning("invoice_pdf_parse_error", error=str(exc))
            return {}

    def _extract_invoice_spreadsheet(self, content: bytes, ext: str) -> dict:
        """Extract invoice data from Excel/CSV template."""
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        data = {}
        for row in ws.iter_rows(values_only=True, max_row=30):
            if row[0] and row[1]:
                key = str(row[0]).lower().strip().replace(" ", "_")
                data[key] = row[1]
        return {
            "invoice_no":   str(data.get("invoice_no", data.get("invoice_number", ""))),
            "party_name":   str(data.get("party_name", data.get("customer_name", ""))),
            "invoice_date": data.get("date", date.today()),
            "subtotal":     float(data.get("subtotal", data.get("net_amount", 0)) or 0),
            "cgst":         float(data.get("cgst", 0) or 0),
            "sgst":         float(data.get("sgst", 0) or 0),
            "igst":         float(data.get("igst", 0) or 0),
            "total":        float(data.get("total", data.get("grand_total", 0)) or 0),
        }

    # ── Row normalisation ─────────────────────────────────────────────────

    def _normalize_bank_row(self, row: dict) -> Optional[dict]:
        """Map various bank CSV column names to standard fields."""
        # Amount detection
        amount = None
        txn_type = "debit"

        # Try debit/credit columns
        debit_keys  = ["debit", "withdrawal", "dr", "debit amount", "withdrawal amt"]
        credit_keys = ["credit", "deposit", "cr", "credit amount", "deposit amt"]

        for k in debit_keys:
            v = self._to_decimal(row.get(k))
            if v and v > 0:
                amount, txn_type = v, "debit"
                break

        if not amount:
            for k in credit_keys:
                v = self._to_decimal(row.get(k))
                if v and v > 0:
                    amount, txn_type = v, "credit"
                    break

        # Try single amount column
        if not amount:
            for k in ["amount", "txn amount", "transaction amount"]:
                v = self._to_decimal(row.get(k))
                if v:
                    amount = abs(v)
                    txn_type = "debit" if v < 0 else "credit"
                    break

        if not amount:
            return None

        # Date detection
        txn_date = None
        for k in ["date", "txn date", "transaction date", "value date", "posting date"]:
            raw = row.get(k)
            if raw:
                txn_date = self._parse_date(str(raw))
                if txn_date:
                    break

        if not txn_date:
            return None

        # Narration
        narration = ""
        for k in ["narration", "description", "particulars", "remarks", "details", "memo"]:
            if row.get(k):
                narration = str(row[k]).strip()
                break

        # Balance
        balance = None
        for k in ["balance", "closing balance", "running balance", "available balance"]:
            v = self._to_decimal(row.get(k))
            if v is not None:
                balance = v
                break

        # Reference
        reference = ""
        for k in ["reference", "ref no", "chq/ref no", "cheque no", "utr", "transaction id"]:
            if row.get(k):
                reference = str(row[k]).strip()
                break

        return {
            "txn_date":  txn_date,
            "amount":    amount,
            "txn_type":  txn_type,
            "narration": narration or "Bank Transaction",
            "reference": reference,
            "balance":   balance,
        }

    def _extract_txns_from_text(self, text: str) -> list[dict]:
        """Regex-based extraction from bank statement text."""
        transactions = []
        # Generic pattern: date + description + amount
        pattern = re.compile(
            r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+'
            r'(.{5,60}?)\s+'
            r'([\d,]+\.?\d*)\s+'
            r'(Dr|Cr|DR|CR)?',
            re.IGNORECASE
        )
        for match in pattern.finditer(text):
            date_str, narration, amount_str, dr_cr = match.groups()
            txn_date = self._parse_date(date_str)
            if not txn_date:
                continue
            amount = self._to_decimal(amount_str)
            if not amount:
                continue
            txn_type = "debit" if (dr_cr or "").upper() == "DR" else "credit"
            transactions.append({
                "txn_date":  txn_date,
                "amount":    amount,
                "txn_type":  txn_type,
                "narration": narration.strip(),
                "reference": "",
                "balance":   None,
            })
        return transactions

    def _parse_invoice_text(self, text: str) -> dict:
        """Extract invoice fields from raw PDF text."""
        def find(patterns: list[str]) -> str:
            for p in patterns:
                m = re.search(p, text, re.IGNORECASE)
                if m:
                    return m.group(1).strip()
            return ""

        def find_amount(patterns: list[str]) -> float:
            for p in patterns:
                m = re.search(p, text, re.IGNORECASE)
                if m:
                    return float(m.group(1).replace(",", ""))
            return 0.0

        return {
            "invoice_no":   find([r"invoice\s*(?:no|number|#)[:\s]+([A-Z0-9/-]+)", r"inv[:\s]+([A-Z0-9/-]+)"]),
            "party_name":   find([r"bill\s*to[:\s]+([^\n]+)", r"customer[:\s]+([^\n]+)", r"party[:\s]+([^\n]+)"]),
            "invoice_date": date.today(),
            "subtotal":     find_amount([r"subtotal[:\s₹]+([0-9,]+\.?\d*)", r"net\s*amount[:\s₹]+([0-9,]+\.?\d*)"]),
            "cgst":         find_amount([r"cgst[:\s₹]+([0-9,]+\.?\d*)"]),
            "sgst":         find_amount([r"sgst[:\s₹]+([0-9,]+\.?\d*)"]),
            "igst":         find_amount([r"igst[:\s₹]+([0-9,]+\.?\d*)"]),
            "total":        find_amount([r"(?:grand\s*)?total[:\s₹]+([0-9,]+\.?\d*)", r"amount\s*due[:\s₹]+([0-9,]+\.?\d*)"]),
        }

    def _to_decimal(self, value) -> Optional[Decimal]:
        if value is None:
            return None
        try:
            cleaned = str(value).replace(",", "").replace("₹", "").replace("Rs", "").strip()
            if not cleaned or cleaned in ("-", ""):
                return None
            return Decimal(cleaned)
        except Exception:
            return None

    def _parse_date(self, date_str: str) -> Optional[date]:
        from datetime import datetime
        formats = [
            "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y",
            "%Y-%m-%d", "%m/%d/%Y", "%d %b %Y", "%d-%b-%Y",
            "%d/%b/%Y", "%d %B %Y",
        ]
        cleaned = date_str.strip()
        for fmt in formats:
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
        return None
